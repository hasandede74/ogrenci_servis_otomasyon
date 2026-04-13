"""
╔══════════════════════════════════════════════════════════════════╗
║     MEBBİS - Taşıma Servis İşlemleri Otomasyonu                 ║
║     Sürüm: 1.0                                                   ║
╚══════════════════════════════════════════════════════════════════╝

Gereksinimler:
    pip install selenium openpyxl

Kullanım:
    python mebbis_otomasyon.py

Excel formatı:
    A Sütunu: Şoför Adı Soyadı
    B Sütunu: Araç Plaka No
    C Sütunu: Öğrenci Adı Soyadı
"""

import time
import sys
import os
from datetime import datetime
from collections import Counter

# ── Bağımlılık kontrolü ──────────────────────────────────────────
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException, StaleElementReferenceException
    )
except ImportError:
    print("=" * 60)
    print("HATA: 'selenium' kurulu degil.")
    print("Lutfen su komutu calistirin:")
    print("    pip install selenium openpyxl")
    print("=" * 60)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("=" * 60)
    print("HATA: 'openpyxl' kurulu degil.")
    print("Lutfen su komutu calistirin:")
    print("    pip install openpyxl")
    print("=" * 60)
    sys.exit(1)

try:
    from tkinter import Tk
    from tkinter.filedialog import askopenfilename
    TKINTER_VAR = True
except ImportError:
    TKINTER_VAR = False


# ══════════════════════════════════════════════════════════════════
#  YARDIMCI FONKSİYONLAR
# ══════════════════════════════════════════════════════════════════

def temizle_metin(metin):
    """Metni buyuk harfe cevirip bastaki/sondaki boslukları temizler."""
    if metin is None:
        return ""
    return str(metin).strip().upper()


def tarih_parse(tarih_str):
    """
    'D.MM.YYYY HH:MM:SS' formatindaki tarihi datetime nesnesine donusturur.
    Ornek: '8.09.2025 00:00:00' -> datetime(2025, 9, 8)
    """
    tarih_str = tarih_str.strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(tarih_str, fmt)
        except ValueError:
            continue
    return None


def dosya_sec():
    """Excel dosyasi secmek icin dialog acar."""
    if TKINTER_VAR:
        root = Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        dosya = askopenfilename(
            title="Ogrenci Listesi Excel Dosyasini Secin",
            filetypes=[("Excel Dosyalari", "*.xlsx *.xls"), ("Tum Dosyalar", "*.*")]
        )
        root.destroy()
        return dosya
    else:
        print("Lutfen dosya yolunu manuel girin:")
        return input("Excel dosya yolu: ").strip()


# ══════════════════════════════════════════════════════════════════
#  EXCEL OKUMA
# ══════════════════════════════════════════════════════════════════

def excel_oku(dosya_yolu):
    """
    Excel dosyasini okuyarak sofor+plaka -> ogrenci listesi sozlugu dondurur.

    Donus:
        ogr_sozluk  : { (sofor_adi, plaka): [ogr_adi1, ogr_adi2, ...] }
        cift_isimler: Ayni isme sahip ogrenci adlarinin listesi
    """
    print(f"\n[Excel Okunuyor] {dosya_yolu}")
    wb = openpyxl.load_workbook(dosya_yolu, read_only=True, data_only=True)
    ws = wb.active

    ogr_sozluk = {}
    tum_ogrenciler = []

    for satir_no, satir in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not satir or (satir[0] is None and satir[1] is None and satir[2] is None):
            continue

        sofor = temizle_metin(satir[0] if len(satir) > 0 else "")
        plaka = temizle_metin(satir[1] if len(satir) > 1 else "")
        ogr   = temizle_metin(satir[2] if len(satir) > 2 else "")

        if not sofor or not plaka or not ogr:
            print(f"  UYARI: Satir {satir_no} eksik veri, atlandi: {satir}")
            continue

        anahtar = (sofor, plaka)
        if anahtar not in ogr_sozluk:
            ogr_sozluk[anahtar] = []
        ogr_sozluk[anahtar].append(ogr)
        tum_ogrenciler.append(ogr)

    wb.close()

    sayac = Counter(tum_ogrenciler)
    cift_isimler = [ad for ad, sayi in sayac.items() if sayi > 1]

    print(f"  => {len(ogr_sozluk)} servis, toplam {len(tum_ogrenciler)} ogrenci kaydi okundu.")
    if cift_isimler:
        print(f"  UYARI - Ayni isimli ogrenciler: {cift_isimler}")

    return ogr_sozluk, cift_isimler


# ══════════════════════════════════════════════════════════════════
#  SELENIUM YARDIMCILARI
# ══════════════════════════════════════════════════════════════════

BEKLEME_SURE = 15
KISA_BEKLE   = 3


def guveli_tikla(driver, element, aciklama="element"):
    """JavaScript ile tiklama."""
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", element)
        time.sleep(0.3)
        driver.execute_script("arguments[0].click();", element)
        return True
    except Exception as e:
        print(f"    UYARI: {aciklama} tiklanamadi: {e}")
        return False


def popup_bekle(driver, zaman_asimi=20):
    """wndOgrenci popup'inin gorunur olmasini bekler."""
    wait = WebDriverWait(driver, zaman_asimi)
    try:
        wait.until(EC.visibility_of_element_located((By.ID, "wndOgrenci_C")))
        time.sleep(0.8)
        return True
    except TimeoutException:
        return False


def popup_kapat_bekle(driver, zaman_asimi=15):
    """Popup kapanmasini bekler."""
    try:
        WebDriverWait(driver, zaman_asimi).until(
            EC.invisibility_of_element_located((By.ID, "wndOgrenci_C"))
        )
        time.sleep(1)
        return True
    except TimeoutException:
        return False


# ══════════════════════════════════════════════════════════════════
#  ANA TABLO OKUMA
# ══════════════════════════════════════════════════════════════════

def ana_tablo_satirlarini_oku(driver):
    """
    rgYaklasikMaliyetler tablosundaki tum satirlari okur.
    Sutun sirasi (HTML'den dogrulandı):
      0: ExpandColumn  1: Plaka  2: SoforAd  3: AracBaslamaTarihi
      4: AracBitisTarihi  5: SoforBaslangic  6: SoforBitis
      7: AktifTasinanOgr  8: ServisOgr  9: Oran  10: DuzenleBtn
    """
    wait = WebDriverWait(driver, BEKLEME_SURE)

    # Once ana tablo elementinin DOM'da olmasini bekle
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located((By.ID, "rgYaklasikMaliyetler_ctl00"))
        )
        print("  Tablo elementi bulundu, satirlar aranıyor...")
    except TimeoutException:
        print("  HATA: rgYaklasikMaliyetler_ctl00 elementi bulunamadi!")
        tables = driver.find_elements(By.CSS_SELECTOR, "table[id]")
        print(f"  Sayfadaki table id'leri: {[t.get_attribute('id') for t in tables[:10]]}")
        return []

    # rgRow satirlarini bekle (30 sn)
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "#rgYaklasikMaliyetler_ctl00 tbody tr.rgRow")
            )
        )
    except TimeoutException:
        print("  UYARI: rgRow bekleme suresi doldu, direkt deneniyor...")

    satirlar = driver.find_elements(
        By.CSS_SELECTOR,
        "#rgYaklasikMaliyetler_ctl00 tbody tr.rgRow,"
        "#rgYaklasikMaliyetler_ctl00 tbody tr.rgAltRow"
    )
    print(f"  Selenium'un buldugu satir sayisi: {len(satirlar)}")

    sonuc = []
    for satir in satirlar:
        try:
            hucreler = satir.find_elements(By.TAG_NAME, "td")
            if len(hucreler) < 9:
                continue

            # Console ciktisinden dogrulanan sutun sirasi:
            # 0:Plaka  1:SoforAd  2:SoforBaslangic  3:SoforBitis
            # 4:ServisBasSaati  5:ServisBitSaati  6:AktifOgr
            # 7:ServisOgr  8:Oran  9:DuzenleBtn
            plaka     = temizle_metin(hucreler[0].text)
            sofor     = temizle_metin(hucreler[1].text)
            sofor_bas = hucreler[2].text.strip()
            sofor_bit = hucreler[3].text.strip()

            # Duzenle butonu: index 9, type="image" input
            duzenle_btn = None
            try:
                duzenle_btn = hucreler[9].find_element(By.TAG_NAME, "input")
            except NoSuchElementException:
                try:
                    duzenle_btn = hucreler[9].find_element(By.TAG_NAME, "a")
                except NoSuchElementException:
                    pass

            if not plaka or not sofor:
                continue

            sonuc.append({
                "plaka": plaka,
                "sofor": sofor,
                "sofor_bas_str": sofor_bas,
                "sofor_bit_str": sofor_bit,
                "sofor_bas": tarih_parse(sofor_bas),
                "sofor_bit": tarih_parse(sofor_bit),
                "duzenle_btn": duzenle_btn,
                "satir_el": satir
            })
        except StaleElementReferenceException:
            continue
        except Exception as e:
            print(f"    UYARI: Satir okunurken hata: {e}")
            continue

    return sonuc


# ══════════════════════════════════════════════════════════════════
#  POPUP İŞLEMİ
# ══════════════════════════════════════════════════════════════════

def popup_ogr_isle(driver, ogr_listesi):
    """
    Acik popuptaki ogrenci grid'ini tarar.
    Excel'deki ogr_listesi ile eslesen, secilebilir (okay_d.png) ogrencileri secer.

    Tik durumlari (HTML'den dogrulandı):
      okay_d.png = Secilebilir (gri tik)  -> TIKLA
      okay_e.png = Zaten secili (yesil)   -> atla
      okay_r.png = Baska aracta (kirmizi) -> atla (disabled)

    Donus: (secilen_sayi, atlanan_sayi, bulunamayan_ogr_listesi)
    """
    wait = WebDriverWait(driver, BEKLEME_SURE)

    try:
        wait.until(EC.presence_of_element_located(
            (By.CSS_SELECTOR,
             "#wndOgrenci_C_rdOgrenciListe_ctl00 tr.rgRow,"
             "#wndOgrenci_C_rdOgrenciListe_ctl00 tr.rgAltRow")
        ))
    except TimeoutException:
        print("    UYARI: Popup ogrenci listesi yuklenemedi!")
        return 0, 0, ogr_listesi

    popup_satirlar = driver.find_elements(
        By.CSS_SELECTOR,
        "#wndOgrenci_C_rdOgrenciListe_ctl00 tr.rgRow,"
        "#wndOgrenci_C_rdOgrenciListe_ctl00 tr.rgAltRow"
    )

    hedef_ogrenciler = set(temizle_metin(o) for o in ogr_listesi)
    bulunan_ogrenciler = set()
    secilen = 0
    atlanan = 0

    for satir in popup_satirlar:
        try:
            hucreler = satir.find_elements(By.TAG_NAME, "td")
            # Popup sutunlari: TC No | Ad | Sinif | Sube | Ogrenci Ekle
            if len(hucreler) < 5:
                continue

            ogr_adi = temizle_metin(hucreler[1].text)

            if ogr_adi not in hedef_ogrenciler:
                continue

            bulunan_ogrenciler.add(ogr_adi)

            # Tik butonunu bul (span[id*='btnSelect'])
            try:
                btn_span = hucreler[4].find_element(
                    By.CSS_SELECTOR, "span[id*='btnSelect']"
                )
            except NoSuchElementException:
                print(f"    UYARI: '{ogr_adi}' icin tik butonu bulunamadi.")
                atlanan += 1
                continue

            stil  = btn_span.get_attribute("style") or ""
            sinif = btn_span.get_attribute("class") or ""

            if "okay_r.png" in stil or "rbDisabled" in sinif:
                print(f"    [X] '{ogr_adi}' -> Baska aracta kayitli, atlaniyor.")
                atlanan += 1
            elif "okay_e.png" in stil:
                print(f"    [V] '{ogr_adi}' -> Zaten secili.")
                secilen += 1
            else:
                # okay_d.png veya bilinmeyen -> tıkla
                print(f"    [>] '{ogr_adi}' seciliyor...")
                guveli_tikla(driver, btn_span, ogr_adi)
                time.sleep(0.4)
                secilen += 1

        except StaleElementReferenceException:
            print(f"    UYARI: Sayfa guncellendi, satir atlandi.")
            continue
        except Exception as e:
            print(f"    UYARI: Satir islenirken hata: {e}")
            atlanan += 1
            continue

    bulunamayan = [o for o in ogr_listesi if temizle_metin(o) not in bulunan_ogrenciler]
    return secilen, atlanan, bulunamayan


def servisi_kaydet(driver):
    """'Servisi Kaydet' butonuna tiklar ve popup'in kapanmasini bekler."""
    wait = WebDriverWait(driver, BEKLEME_SURE)
    try:
        kaydet_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "span[id='wndOgrenci_C_btnSave']")
        ))
        guveli_tikla(driver, kaydet_btn, "Servisi Kaydet")
        print("    [Kaydet] Servisi Kaydet tiklandi, bekleniyor...")
        time.sleep(1.5)

        # SweetAlert onay kutusu gelebilir
        try:
            confirm_btn = driver.find_element(
                By.CSS_SELECTOR, ".swal2-confirm"
            )
            if confirm_btn.is_displayed():
                confirm_btn.click()
                print("    [OK] SweetAlert onaylandi.")
                time.sleep(1)
        except NoSuchElementException:
            pass

        # SweetAlert sonrasi popup kapanmaz, direkt X butonuyla kapat
        time.sleep(0.5)
        try:
            close_btn = driver.find_element(By.CSS_SELECTOR, ".rwCloseButton")
            guveli_tikla(driver, close_btn, "X Kapat butonu")
            print("    [OK] Popup X butonu ile kapatildi.")
            time.sleep(1)
        except NoSuchElementException:
            # X butonu yoksa kapanmasini bekle
            kapandi = popup_kapat_bekle(driver)
            if kapandi:
                print("    [OK] Popup kapandi.")
            else:
                print("    UYARI: Popup kapatilamadi!")
        return True

    except TimeoutException:
        print("    HATA: Servisi Kaydet butonu bulunamadi!")
        return False


# ══════════════════════════════════════════════════════════════════
#  ANA OTOMASYON DONGUSU
# ══════════════════════════════════════════════════════════════════

def otomasyon_calistir(driver, ogr_sozluk):
    """Ana otomasyon dongusu."""
    bugun = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    print(f"\nBugunun tarihi: {bugun.strftime('%d.%m.%Y')}")

    # Rapor yapisi
    rapor = {
        "servis_raporu": [],      # Her servis icin {servis, durum, aciklama}
        "ogrenci_raporu": {},     # {ogr_adi: {servis, durum}}
        "hatalar": []
    }

    # Excel'deki tum ogrencileri "islenmedi" olarak baslat
    for (e_sofor, e_plaka), ogrenciler in ogr_sozluk.items():
        for ogr in ogrenciler:
            rapor["ogrenci_raporu"][ogr] = {"servis": f"{e_sofor} / {e_plaka}", "durum": "Yerlestirilmedi"}

    # Hangi servisler zaten islendi (tekrar islemeyi onlemek icin)
    islenen_servisler = set()

    while True:
        print("\n--- Ana tablo okunuyor... ---")
        try:
            satirlar = ana_tablo_satirlarini_oku(driver)
        except Exception as e:
            print(f"HATA: Tablo okunamadi: {e}")
            rapor["hatalar"].append(f"Tablo okunamadi: {e}")
            break

        if not satirlar:
            print("Tabloda islenecek satir bulunamadi.")
            break

        print(f"Toplam {len(satirlar)} satir bulundu.")

        # Excel'de islenmesi gereken servisler hepsi bitti mi?
        bekleyen_servisler = []
        for (e_sofor, e_plaka) in ogr_sozluk.keys():
            anahtar = f"{e_sofor} / {e_plaka}"
            if anahtar not in islenen_servisler:
                bekleyen_servisler.append(anahtar)

        if not bekleyen_servisler:
            print("Excel'deki tum servisler islendi. Otomasyon tamamlandi.")
            break

        isleme_yapildi = False

        for idx, satir in enumerate(satirlar):
            plaka      = satir["plaka"]
            sofor      = satir["sofor"]
            sofor_bas  = satir["sofor_bas"]
            sofor_bit  = satir["sofor_bit"]
            servis_key = f"{sofor} / {plaka}"

            # Zaten islendiyse atla
            if servis_key in islenen_servisler:
                continue

            # --- Tarih kontrolu ---
            if sofor_bas is None or sofor_bit is None:
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "ATLANDI",
                    "aciklama": f"Tarih parse edilemedi ({satir['sofor_bas_str']} - {satir['sofor_bit_str']})"
                })
                islenen_servisler.add(servis_key)
                continue

            if not (sofor_bas <= bugun <= sofor_bit):
                # Sadece bu site satirini islenmiş say, Excel anahtarini ekleme!
                # Boylece ayni soforun gecerli tarihli satiri atlanmaz.
                islenen_servisler.add(servis_key)
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "ATLANDI",
                    "aciklama": f"Tarih araligi disinda ({satir['sofor_bas_str']} - {satir['sofor_bit_str']})"
                })
                print(f"  [{idx+1}] {servis_key}: Tarih araligi disinda, atlandi.")
                continue

            # --- Excel eslesmesi: once plaka+sofor, sonra sadece plaka ---
            ogr_listesi = None
            eslesme_key = None
            for (e_sofor, e_plaka), ogrenciler in ogr_sozluk.items():
                if e_sofor == sofor and e_plaka == plaka:
                    ogr_listesi = ogrenciler
                    eslesme_key = f"{e_sofor} / {e_plaka}"
                    break

            if ogr_listesi is None:
                for (e_sofor, e_plaka), ogrenciler in ogr_sozluk.items():
                    if e_plaka == plaka and f"{e_sofor} / {e_plaka}" not in islenen_servisler:
                        ogr_listesi = ogrenciler
                        eslesme_key = f"{e_sofor} / {e_plaka}"
                        print(f"  BILGI: Plaka ile eslesti - Excel='{e_sofor}', Site='{sofor}'")
                        break

            if ogr_listesi is None:
                # Bu site satirinda Excel eslesmesi yok, sadece bu satiri atla
                islenen_servisler.add(servis_key)
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "ATLANDI",
                    "aciklama": "Excel'de bu servis icin ogrenci bulunamadi"
                })
                print(f"  [{idx+1}] {servis_key}: Excel'de bulunamadi, atlandi.")
                continue

            # Bu Excel servisi zaten islendiyse bu site satirini da atla
            if eslesme_key in islenen_servisler:
                islenen_servisler.add(servis_key)
                continue

            # --- Duzenle butonuna tikla ---
            print(f"\n  [{idx+1}] {servis_key} isleniyor...")
            duzenle_btn = satir["duzenle_btn"]
            if duzenle_btn is None:
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "HATA",
                    "aciklama": "Duzenle butonu bulunamadi"
                })
                rapor["hatalar"].append(f"{servis_key}: Duzenle butonu yok")
                islenen_servisler.add(servis_key)
                continue

            try:
                guveli_tikla(driver, duzenle_btn, "Duzenle butonu")
            except Exception as e:
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "HATA",
                    "aciklama": f"Tiklama hatasi: {e}"
                })
                rapor["hatalar"].append(f"{servis_key}: Tiklama hatasi - {e}")
                islenen_servisler.add(servis_key)
                continue

            # --- Popup bekle ---
            print(f"    Popup yukleniyor...")
            if not popup_bekle(driver):
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "HATA",
                    "aciklama": "Popup acilmadi"
                })
                rapor["hatalar"].append(f"{servis_key}: Popup acilmadi")
                islenen_servisler.add(servis_key)
                continue

            # --- Popup bilgilerini logla ---
            try:
                popup_sofor = driver.find_element(By.ID, "wndOgrenci_C_txtSoforAd").get_attribute("value")
                popup_plaka = driver.find_element(By.ID, "wndOgrenci_C_txtPlaka").get_attribute("value")
                print(f"    Popup: Sofor='{popup_sofor}', Plaka='{popup_plaka}'")
            except Exception:
                pass

            # --- Ogrencileri isle ---
            secilen, atlanan_ogr, bulunamayan = popup_ogr_isle(driver, ogr_listesi)
            print(f"    Sonuc: Secilen={secilen}, Atlanan={atlanan_ogr}, Bulunamayan={len(bulunamayan)}")

            # Ogrenci raporunu guncelle
            for ogr in ogr_listesi:
                ogr_temiz = temizle_metin(ogr)
                if ogr_temiz not in [temizle_metin(b) for b in bulunamayan]:
                    rapor["ogrenci_raporu"][ogr] = {
                        "servis": servis_key,
                        "durum": "Servise yerlestirildi"
                    }
                else:
                    rapor["ogrenci_raporu"][ogr] = {
                        "servis": servis_key,
                        "durum": "Popup'ta bulunamadi - Manuel giris gerekli"
                    }

            # --- Kaydet ---
            kayit_ok = servisi_kaydet(driver)

            if kayit_ok:
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "TAMAMLANDI",
                    "aciklama": f"{secilen} ogrenci yerlestirildi"
                })
                islenen_servisler.add(servis_key)
                if eslesme_key:
                    islenen_servisler.add(eslesme_key)
                isleme_yapildi = True
            else:
                rapor["servis_raporu"].append({
                    "servis": servis_key,
                    "durum": "HATA",
                    "aciklama": "Kaydetme basarisiz"
                })
                rapor["hatalar"].append(f"{servis_key}: Kaydetme basarisiz")
                islenen_servisler.add(servis_key)

            time.sleep(2)
            break  # Tablo yenilendigi icin donguyu kir, while devam eder

        if not isleme_yapildi:
            # Hic islem yapilmadi, tum satirlar tarandı -> bitti
            break

    return rapor


# ══════════════════════════════════════════════════════════════════
#  RAPOR
# ══════════════════════════════════════════════════════════════════

def rapor_yazdir(rapor, cift_isimler):
    """Otomasyon sonuc raporunu konsola yazdirir."""
    print("\n")
    print("=" * 65)
    print("                   OTOMASYON RAPORU")
    print("=" * 65)

    # --- SERVİS RAPORU ---
    print("\n[SERVİS RAPORU]")
    print("-" * 65)
    for r in rapor["servis_raporu"]:
        durum   = r["durum"]
        servis  = r["servis"]
        aciklama = r.get("aciklama", "")
        if durum == "TAMAMLANDI":
            isaret = "[OK]"
        elif durum == "ATLANDI":
            isaret = "[--]"
        else:
            isaret = "[!!]"
        print(f"  {isaret} {servis}")
        if aciklama:
            print(f"       -> {aciklama}")

    if not rapor["servis_raporu"]:
        print("  (islem yapilmadi)")

    # --- ÖĞRENCİ RAPORU ---
    print("\n[OGRENCİ RAPORU]")
    print("-" * 65)
    for ogr, bilgi in rapor["ogrenci_raporu"].items():
        durum  = bilgi["durum"]
        servis = bilgi["servis"]
        if "yerlestirildi" in durum.lower():
            isaret = "[OK]"
        else:
            isaret = "[!!]"
        print(f"  {isaret} {ogr}")
        print(f"       -> {durum} ({servis})")

    if not rapor["ogrenci_raporu"]:
        print("  (ogrenci bilgisi yok)")

    # --- HATALAR ---
    if rapor["hatalar"]:
        print(f"\n[HATALAR]")
        print("-" * 65)
        for h in rapor["hatalar"]:
            print(f"  [!!] {h}")

    # --- AYNI İSİMLİ ÖĞRENCİLER ---
    if cift_isimler:
        print(f"\n[UYARI - MANUEL KONTROL GEREKLİ]")
        print("-" * 65)
        print("  Asagidaki ogrenciler excelde birden fazla kez gecmektedir.")
        print("  Dogru servise atandiklarini lutfen manuel dogrulayin:")
        for ad in cift_isimler:
            print(f"  [!!] {ad}")

    print("\n" + "=" * 65)
    print("Program tamamlandi.")
    print("=" * 65)



def rapor_kaydet(rapor, cift_isimler, excel_yolu):
    """Raporu hem TXT hem Excel dosyasi olarak kaydeder."""
    import os
    zaman = datetime.now().strftime("%Y%m%d_%H%M%S")
    klasor = os.path.dirname(os.path.abspath(excel_yolu))
    txt_dosya  = os.path.join(klasor, f"mebbis_rapor_{zaman}.txt")
    xlsx_dosya = os.path.join(klasor, f"mebbis_rapor_{zaman}.xlsx")

    # --- TXT ---
    satirlar = []
    satirlar.append("=" * 65)
    satirlar.append("              MEBBiS OTOMASYON RAPORU")
    satirlar.append(f"              {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    satirlar.append("=" * 65)

    satirlar.append("\n[SERViS RAPORU]")
    satirlar.append("-" * 65)
    for r in rapor["servis_raporu"]:
        isaret = "[OK]" if r["durum"] == "TAMAMLANDI" else ("[--]" if r["durum"] == "ATLANDI" else "[!!]")
        satirlar.append(f"  {isaret} {r['servis']}")
        if r.get("aciklama"):
            satirlar.append(f"       -> {r['aciklama']}")

    satirlar.append("\n[OGRENCi RAPORU]")
    satirlar.append("-" * 65)
    for ogr, bilgi in rapor["ogrenci_raporu"].items():
        isaret = "[OK]" if "yerlestirildi" in bilgi["durum"].lower() else "[!!]"
        satirlar.append(f"  {isaret} {ogr}")
        satirlar.append(f"       -> {bilgi['durum']} ({bilgi['servis']})")

    if rapor["hatalar"]:
        satirlar.append("\n[HATALAR]")
        satirlar.append("-" * 65)
        for h in rapor["hatalar"]:
            satirlar.append(f"  [!!] {h}")

    if cift_isimler:
        satirlar.append("\n[UYARI - AYNI iSiMLi OGRENCiLER]")
        satirlar.append("-" * 65)
        for ad in cift_isimler:
            satirlar.append(f"  [!!] {ad}")

    satirlar.append("\n" + "=" * 65)

    try:
        with open(txt_dosya, "w", encoding="utf-8") as f:
            f.write("\n".join(satirlar))
        print(f"  TXT raporu kaydedildi: {txt_dosya}")
    except Exception as e:
        print(f"  UYARI: TXT kaydedilemedi: {e}")

    # --- EXCEL ---
    try:
        wb = openpyxl.Workbook()

        # Sayfa 1: Servis Raporu
        ws1 = wb.active
        ws1.title = "Servis Raporu"
        ws1.append(["Servis", "Durum", "Aciklama"])
        for r in rapor["servis_raporu"]:
            ws1.append([r["servis"], r["durum"], r.get("aciklama", "")])

        # Sayfa 2: Ogrenci Raporu
        ws2 = wb.create_sheet("Ogrenci Raporu")
        ws2.append(["Ogrenci Adi", "Durum", "Servis"])
        for ogr, bilgi in rapor["ogrenci_raporu"].items():
            ws2.append([ogr, bilgi["durum"], bilgi["servis"]])

        # Sayfa 3: Hatalar
        if rapor["hatalar"] or cift_isimler:
            ws3 = wb.create_sheet("Uyarilar")
            ws3.append(["Tur", "Aciklama"])
            for h in rapor["hatalar"]:
                ws3.append(["HATA", h])
            for ad in cift_isimler:
                ws3.append(["AYNI iSiM", ad])

        # Sutun genisliklerini ayarla
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(xlsx_dosya)
        print(f"  Excel raporu kaydedildi: {xlsx_dosya}")
    except Exception as e:
        print(f"  UYARI: Excel kaydedilemedi: {e}")

# ══════════════════════════════════════════════════════════════════
#  GİRİŞ NOKTASI
# ══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("  MEBBiS Tasima Servis Otomasyonu")
    print("=" * 60)
    print("\nHos geldiniz!\n")

    # 1. Excel sec
    print("[ADIM 1/4] Ogrenci listesi Excel dosyasini secin...")
    excel_yolu = dosya_sec()
    if not excel_yolu:
        print("Dosya secilmedi. Program sonlandiriliyor.")
        sys.exit(0)

    # 2. Excel oku
    try:
        ogr_sozluk, cift_isimler = excel_oku(excel_yolu)
    except Exception as e:
        print(f"HATA: Excel okunurken hata: {e}")
        sys.exit(1)

    if not ogr_sozluk:
        print("HATA: Excel'de gecerli veri bulunamadi.")
        sys.exit(1)

    # 3. Chrome ac
    print("\n[ADIM 2/4] Chrome acilacak.")
    input("  ENTER'a basin...")

    chrome_options = Options()
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--lang=tr")

    try:
        driver = webdriver.Chrome(options=chrome_options)
    except Exception as e:
        print(f"HATA: Chrome acilamadi: {e}")
        print("ChromeDriver'in kurulu ve PATH'te oldugunden emin olun.")
        print("Indir: https://googlechromelabs.github.io/chrome-for-testing/")
        sys.exit(1)

    driver.get("https://mebbis.meb.gov.tr")

    # 4. Manuel giris
    print("\n[ADIM 3/4] Manuel giris:")
    print("  1. MEBBiS'e giris yapin.")
    print("  2. 'Tasima Servis Islemleri' sayfasina gidin.")
    print("  3. 'Sorgula' butonuna basin ve liste yuklensin.")
    print("  4. Hazir olunca asagida ENTER'a basin.")
    input("\n  ENTER'a basarak otomasyonu baslatin...")


    # Tum sekmeleri tara, tasimali.meb.gov.tr olan sekmeye gec
    print("\n  Aktif sekme aranıyor...")
    hedef_handle = None
    for handle in driver.window_handles:
        driver.switch_to.window(handle)
        url = driver.current_url
        baslik = driver.title
        print(f"  Sekme: {baslik} | URL: {url}")
        if "tasimali.meb.gov.tr" in url or "ETS01025" in url:
            hedef_handle = handle
            print(f"  => Hedef sekme bulundu!")
            break

    if hedef_handle is None:
        driver.switch_to.window(driver.window_handles[-1])
        print(f"  => Hedef sekme bulunamadi, son sekme: {driver.current_url}")

    time.sleep(2)

    # 5. Otomasyon
    print("\n[ADIM 4/4] Otomasyon basliyor...\n")
    try:
        rapor = otomasyon_calistir(driver, ogr_sozluk)
    except KeyboardInterrupt:
        print("\nKullanici tarafindan durduruldu.")
        rapor = {
            "islenen": [], "atlanan_tarih": [], "atlanan_excel_yok": [],
            "bulunamayan_ogrenciler": {}, "hatalar": ["Kullanici durdurdu"]
        }
    except Exception as e:
        print(f"\nBeklenmedik hata: {e}")
        import traceback
        traceback.print_exc()
        rapor = {
            "islenen": [], "atlanan_tarih": [], "atlanan_excel_yok": [],
            "bulunamayan_ogrenciler": {}, "hatalar": [str(e)]
        }

    # 6. Rapor
    rapor_yazdir(rapor, cift_isimler)
    rapor_kaydet(rapor, cift_isimler, excel_yolu)

    input("\nKapatmak icin ENTER'a basin...")
    try:
        driver.quit()
    except Exception:
        pass


if __name__ == "__main__":
    main()
